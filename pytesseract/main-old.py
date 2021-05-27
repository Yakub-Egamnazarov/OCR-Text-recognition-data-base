from pprint import pprint
import os 
import cv2
import pytesseract
from pytesseract import Output
from PIL import Image
import pyheif
import openpyxl



# opening heic files 
def heic_to_pil(input_abs):
    heic_file = pyheif.read(input_abs)
    img = Image.frombytes(heic_file.mode, 
                          heic_file.size, 
                          heic_file.data, 
                          'raw', 
                          heic_file.mode, 
                          heic_file.stride,)
    
    return img 

# cropping the images 
def img_cropping(input_dir, output_dir):
    img_names = os.listdir(input_dir)
    
    count = 0
    for img in img_names: 
        img_obj = heic_to_pil(os.path.join(input_dir, img))
        img_cropped = img_obj.crop((0, 
                                    0, 
                                    img_obj.size[0] * 4 // 5, 
                                    img_obj.size[1] // 3))
        img_cropped.save(os.path.join(output_dir, 
                            img.split('.')[0] + '-cropped.jpeg'))
        count += 1
        print('done crpping: ', count)


# function reads each picture inside the folder and writes 
# to text files
def img_to_text_2(input_dir):
    img_names = os.listdir(input_dir)
    img_names.sort()
    data = {}
    
    for count, img in enumerate(img_names):
        img_path = os.path.join(input_dir, img)
        
        # STEP-1: Reading the image with OpenCV
        img_obj = cv2.imread(img_path)
        
        # STEP-2: converting the image into the gray scale
        img_obj_gray = cv2.cvtColor(img_obj, cv2.COLOR_BGRA2GRAY)
        
        # STEP-3: performing OTSU threshold 
        img_obj_threshold = cv2.threshold(img_obj_gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        
        # STEP-4: configuring the parameters for tesseract
        custom_config = r'--oem 3 --psm 6'
        
        # STEP-5: feeding image to pytesseract
        details = pytesseract.image_to_data(img_obj_threshold,output_type=Output.DICT, config=custom_config, lang='eng')
        
        # STEP-6: Drawing the bounding boxes for the image
        total_boxes = len(details['text'])
        
        for seq_num in range(total_boxes):
            if int(details['conf'][seq_num]) > 30:
                (x, y, w, h) = (details['left'][seq_num], details['top'][seq_num], details['width'][seq_num], details['height'][seq_num])
                img_obj_boxed = cv2.rectangle(img_obj, (x-5, y-5), (x+w+5, y+h+5), (0, 255, 0), 1)
                
        # STEP-7: Saving the boxed image
        cv2.imwrite(os.path.join(os.getcwd(), 'img_boxed', img.split('-')[0] + '-boxed.jpg'), img_obj_boxed)
        
        # STEP-8: Parsing the text 
        parse_txt, word_list, last_word = [], [], ''
        for word in details['text']:
            if word != '':
                word_list.append(word)
                last_word = word
            if (last_word != '' and word == '') or (word == details['text'][-1]):
                parse_txt.append(word_list)
                word_list = []
        
        # print('=> Parsed text:') TODO
        # pprint(parse_txt)  TODO
        print('=> done OCR:', count)
        data[img] = parse_txt

    return data

def img_to_text(input_dir, output_dir):
    img_names = os.listdir(input_dir)
    img_names.sort()

    count = 0
    for img in img_names:
        img_path = os.path.join(input_dir, img)
        
        # Read the image from which text needs to be extracted
        img_obj = cv2.imread(img_path)
        
        # Preprocessing the image
        # 1. Convert the image to gray scale
        gray_img = cv2.cvtColor(img_obj, cv2.COLOR_BGR2GRAY)
        
        # 2. Performing OTSU Threshold
        ret, thresh1 = cv2.threshold(gray_img, 0, 255, 
                            cv2.THRESH_OTSU | cv2.THRESH_BINARY_INV)
        
        # Specify structure shape and kernel size 
        # Kernel size increases or decreases the area 
        # of the rectangle to be detected. 
        # A smaller value like (10, 10) will detect 
        # each word instead of a sentence. 
        rect_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, 
                                                (50, 20))
        
        # applying dilation on the threshold image
        dilation = cv2.dilate(thresh1, rect_kernel, iterations=1)
        
        # Finding contours
        contours, hierarchy = cv2.findContours(dilation, 
                                               cv2.RETR_EXTERNAL, 
                                               cv2.CHAIN_APPROX_NONE)
        
        # creating a copy of the image 
        img_obj_2 = img_obj.copy()
        
        # a text file created for the image 
        
        # Looping through the identified contours 
        # Then Rectangular part i scropped and passed
        # on to pytesseract for extracting text from it 
        # Extracted text is then written in to the file 
        text_all= ''
        for cnt in contours: 
            x, y, w, h = cv2.boundingRect(cnt)
            
            # Drawing a rectangle on copied image 
            rect = cv2.rectangle(img_obj_2, 
                                 (x, y), 
                                 (x + w, y + h), 
                                 (0, 255, 0), 
                                 2)
            
            # cropping the text block for giving input to ocr 
            cropped = img_obj_2[y: y + h, x: x + w]
            text = pytesseract.image_to_string(cropped)
            text_all += text.strip() + '\n'

        
        file_path = os.path.join(output_dir, 
                                 img.split('-')[0] + '-raw.txt')
        file = open(file_path, 'w')
        file.write(text_all)
        file.close()
        
        cv2.imwrite(os.path.join(os.getcwd(), 
                                 'img_boxed', 
                                 img.split('-')[0] + '-boxed.jpg'), 
                    img_obj_2)
        count += 1
        print ('done ocr: ', count)
        
        
def cutting_list(lst):
    if 'CONTACT DETAILS\n' in lst:
        cutting_index = lst.index('CONTACT DETAILS\n')
        
        return lst[0: cutting_index]
    else:
        return lst

def clear_empty_el(lst):
    new_list = []
    for item in lst:
        if len(item) > 1 and item != '\n':
            new_list.append(item.strip())

    for item in new_list:
        if (item.startswith('w') or item.startswith('W')):
            break
        else:
            new_list.remove(item)
    print('+++')
    pprint(new_list)
    print('+++')
    return new_list


def extract_website(lst):
    website = ''
    if 'Website' in lst:
        lst.remove('Website')
    
    for item in lst:
        if item.startswith('www') or item.startswith('Www'):
            website = item
            lst.remove(item)

    if website.startswith('W'):        
        website = website[1:-1]
        
    return lst, website

def extract_email(lst):    
    if 'Email' in lst:
        lst.remove('Email')
    if 'email' in lst: 
        lst.remove('email')
        
    email = lst[0]
    lst.remove(lst[0])
    
    return lst, email


def extract_things(lst, key1, key2):
    elem = ''
    if key1 in lst:
        lst.remove(key1)
        elem = lst[0]
        lst.remove(lst[0])
    if key2 in lst: 
        lst.remove(key2)
        elem = lst[0]
        lst.remove(lst[0])
    return lst, elem

def extract_company(lst):
    comp_name = ''
    if 'Company Name' in lst:
        lst.remove('Company Name')
        comp_name = lst[-1]
        lst.remove(lst[-1])
    if 'company name' in lst:
        lst.rmove('company name')
        comp_name = lst[-1]
        lst.remove(lst[-1])
        
    return lst, comp_name

def extract_address(lst):
    address = ''
    if 'Address' in lst:
        lst.remove('Address')
        if len(lst) > 1:
            lst.reverse()
        address = ' '.join(lst)

    if 'address' in lst:
        lst.remove('address')
        if len(lst) > 1:
            lst.reverse()
        address = ' '.join(lst)
    return lst, address

# function searches for the country in the address
def find_country(string):
    # country = ''
    lst = string.split(' ')
    if lst[-1].isdecimal():
        lst.remove(lst[-1])
    
    country = lst[-1]
    
    return country


# function extracts the position and delegate name from the delegate name 
def extract_position(string):
    lst = string.split(',')
    position = lst[-1].strip()
    delegate_name = lst[0]
    
    return position, delegate_name
    
# function corrects the website 
def correct_website(string):
    website = []
    
    lst = string.split(';')
    for item in lst: 
        i = [el.strip() for el in item.split('.')]
        web = '.'.join(i)
        website.append(web)

    return website 


def correct_website_2(websites, emails):
    new_websites = []
    print(websites)
    print(emails)
    for website in websites: 
        for email in emails:
            print(email)
            web_base = website.split('www.')[-1]
            email_stucture = email.split(web_base)
            if email_stucture[-1] != '':
                print('> correcting website...')
                new_websites.append('www.' + web_base + email_stucture[-1])
            else: 
                new_websites.append(website)
    print('> Corrected website: ', new_websites)    
    return new_websites

# making email from website base and email base
def cleaning_email(web_base, email):
    print('< Cleaning EMAIL...')
    email_stucture = email.split(web_base)
    if email_stucture[-1] != '':
        print('The website should be corrected')

    if email_stucture[0].startswith('('):
        email_stucture[0] = 'l' + email_stucture[0][1:-1]
        print('>>>l added to the begginning')
    
    if email_stucture[0].endswith('ld'):
        email_stucture[0] = email_stucture[0][0:-2] + '@'
        print('>>>"ld" deleted')
    elif email_stucture[0].endswith('id'):
        email_stucture[0] = email_stucture[0][0:-2] + '@'
        print('>>>"id" deleted')
    elif email_stucture[0].endswith('fd'):
        email_stucture[0] = email_stucture[0][0:-2] + '@'
        print('>>>"fd" deleted')
    elif email_stucture[0].endswith('(d'):
        email_stucture[0] = email_stucture[0][0:-2] + '@'
        print('>>>"(d" deleted')
    elif email_stucture[0].endswith('{d'):
        email_stucture[0] = email_stucture[0][0:-2] + '@'
        print('>>>"{d" deleted')
    elif email_stucture[0].endswith('d'):
        email_stucture[0] = email_stucture[0][0:-1] + '@'
        print('>>>"d" deleted')
    elif email_stucture[0].endswith('('):
        email_stucture[0] = email_stucture[0][0:-1] + '@'
        print('>>>"(" deleted')
    elif email_stucture[0].endswith('{'):
        email_stucture[0] = email_stucture[0][0:-1] + '@'
        print('>>>"{" deleted')
    
    email = web_base.join(email_stucture)
    # print(email)
    return email


# function makes a correction to the email from the website reference
def correct_email(email_raw, website_lst):
    
    print('>Correcting EMAIL...')
    if len(email_raw) > 1:
        email_raw_lst = [el.strip() for el in email_raw.split(';')]
        
        email_clean = []
        
        for website in website_lst:
            web_base = website.split('www')[-1].strip('.')
            for email in email_raw_lst:
                if web_base in email and web_base != '':
                    email_clean.append(cleaning_email(web_base, email))
                    
        
        print(email_clean)


    
        return email_clean
    else: 
        return email_raw

def parse_list(lst):
    data = {}
    
    lst, data['website'] = extract_website(lst)
    lst, data['email'] = extract_email(lst)
    lst, data['fax'] = extract_things(lst, 'Fax', 'fax')
    lst, data['tel'] = extract_things(lst, 'Tel', 'tel')
    lst, data['delegate_name'] = extract_things(lst, 'Delegate Name', 
                                                'delegate name')
    lst, data['company_name'] = extract_company(lst)
    lst, data['address'] = extract_address(lst)
    
    # Finding the country from the address
    data['country'] = find_country(data['address'])
    
    # Dividing the delegate name from the position:
    data['position'], data['delegate_name'] = extract_position(data['delegate_name'])
    
    # correcting the website 
    data['website-2'] = correct_website(data['website'])
    
    # correcting the email
    data['email-2'] = correct_email(data['email'], data['website-2'])
    
    # correcting the webstie second time 
    data['website-2'] = correct_website_2(data['website-2'], data['email-2'])
    
    print()
    return data





# extracting data from raw text files 
def text_list(input_dir):
    txt_names = os.listdir(input_dir) # extracting the file names
    txt_names.sort()    # sorting the file names

    
    data = []
    for count, txt in enumerate(txt_names):
        txt_path = os.path.join(input_dir, txt)
        file = open(txt_path, 'r') 
        
        # count += 1
        print('file - {n}: {name}'.format(n=count, name=txt))
        
        file_list = clear_empty_el(cutting_list(file.readlines()))
        # file_list = clear_empty_el(file_list)
        file_data = parse_list(file_list)
        file_data['img_name'] = txt.split('.')[0]
        print()
        print('> List parsed.')
        for item in file_data:
            print('{}: {}'.format(item, file_data[item]))

        data.append(file_data)
        print()
        print('--- --- --- --- --- --- ')

    return data

# function writes extracted data to dababase excel sheet
def write_to_DB(data_in, out_dir):
    # A1 = No
    # B1 = Company Name 
    # C1 = Country 
    # D1 = Website
    # E1 = Delegate Name 
    # F1 = Position 
    # G1 = Email
    # H1 = Tel 
    # I1 = Fax 
    # J1 = Img-file 
    # K1 = Company introduction / Profile
    
    # TODO read the DB file from the folder 
    file_path = os.path.join(out_dir, 'customer_list-0.xlsx')
    wb_out = openpyxl.load_workbook(file_path)
    
    # TODO get access to the working sheet
    cur_sheet = wb_out['English']
    
    # TODO Check whether rows are empty 
    max_rows = cur_sheet.max_row

    print('The maximum rows: ', max_rows)
    start_row = 2

    # TODO start deliver the data to a certain row from the empty row 
    for count, data in enumerate(data_in):
        # Each data goes to one row in the sheet
        count_abs = count + start_row
        
        cur_sheet.cell(row=count_abs, column=1).value = count_abs - 1
        cur_sheet.cell(row=count_abs, column=2).value = data['company_name']
        cur_sheet.cell(row=count_abs, column=3).value = data['country']
        cur_sheet.cell(row=count_abs, column=4).value = ';'.join(data['website-2'])
        cur_sheet.cell(row=count_abs, column=5).value = data['delegate_name']
        cur_sheet.cell(row=count_abs, column=6).value = data['position']
        cur_sheet.cell(row=count_abs, column=7).value = ';'.join(data['email-2'])
        cur_sheet.cell(row=count_abs, column=8).value = data['tel']
        cur_sheet.cell(row=count_abs, column=9).value = data['fax']
        cur_sheet.cell(row=count_abs, column=10).value = data['img_name']


    
    # TODO after completing save the updated new worksheet 
    wb_out.save(file_path)
    
    
    
    
    print('>Completed writing to DataBase')


# ==========================================================
# ==========================================================
cwd = os.getcwd()

cur_dir = os.path.join(cwd, 'pytesseract')
img_heic_dir = os.path.join(cwd, 'img_heic')
img_jpg_dir = os.path.join(cwd, 'img_jpg')
img_cropp_dir = os.path.join(cwd, 'img_cropped')
out_txt_dir = os.path.join(cwd, 'output-txt')
out_data_dir = os.path.join(cwd, 'output_data')
print(os.path.exists(out_data_dir))

# converting to  jpeg file format from .heic apple image format
# and cropping Image object
img_cropping(img_heic_dir, img_cropp_dir) 

# reading the text from cropped files
print('ocr processing...')
# DATA = img_to_text_2(img_cropp_dir)
# img_to_text(img_cropp_dir, out_txt_dir) TODO
print('--- --- --- OCR-done --- --- ---')

for key in DATA:
    print('--- key =>', key)
    print(DATA[key])

# print('reading data...')  TODO
DATA = text_list(out_txt_dir) TODO  
# print('--- --- --- reading done! --- --- --- ') TODO
# print() TODO

# print('writing to Excel DB...') TODO
# write_to_DB(DATA, out_data_dir) TODO
# print('--- --- --- Writing done! --- --- --- ') TODO
# print() TODO

print('--- --- --- ALL TASKS COMPLETED --- --- --- ')