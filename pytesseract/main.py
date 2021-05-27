from pprint import pprint
import csv
import os 
import cv2
import pytesseract
from pytesseract import Output
from PIL import Image
import pyheif
import openpyxl
import re

# ==========================================================
# --- --- FUNCTIONS --- ---
# ==========================================================

# opening heic files 
def heic_to_pil(input_abs):
    heic_file = pyheif.read(input_abs)
    img = Image.frombytes(heic_file.mode, heic_file.size, heic_file.data, 'raw', heic_file.mode, heic_file.stride,)
    
    return img 


# cropping the images 
def img_cropping(input_dir, output_dir):
    img_names = os.listdir(input_dir)
    
    for count, img in enumerate(img_names): 
        img_obj = heic_to_pil(os.path.join(input_dir, img))
        img_cropped = img_obj.crop((0, 
                                    0, 
                                    img_obj.size[0] * 4 // 5, 
                                    img_obj.size[1] // 3))
        img_cropped.save(os.path.join(output_dir, 
                            img.split('.')[0] + '-cropped.jpeg'))
        print('done crpping: ', count)

# function reads each picture inside the folder and writes 
# to text files
def img_to_text(input_dir, output_dir):
    img_names = os.listdir(input_dir)
    img_names.sort()
    data = {}
    
    for count, img in enumerate(img_names):
        img_path = os.path.join(input_dir, img)
        
        # STEP-1: Reading the image with OpenCV
        img_obj = cv2.imread(img_path)
        
        # STEP-2: converting the image into the gray scale
        img_obj_gray = cv2.cvtColor(img_obj, cv2.COLOR_BGRA2GRAY)
        
        # STEP-3: performing OTSU threshold 
        img_obj_threshold = cv2.threshold(img_obj_gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        
        # STEP-4: configuring the parameters for tesseract
        custom_config = r'--oem 3 --psm 6'
        
        # STEP-5: feeding image to pytesseract
        details = pytesseract.image_to_data(img_obj_threshold,output_type=Output.DICT, config=custom_config, lang='eng')
        
        # STEP-6: Drawing the bounding boxes for the image
        total_boxes = len(details['text'])
        
        for seq_num in range(total_boxes):
            if int(details['conf'][seq_num]) > 30:
                (x, y, w, h) = (details['left'][seq_num], details['top'][seq_num], details['width'][seq_num], details['height'][seq_num])
                img_obj_boxed = cv2.rectangle(img_obj, (x-5, y-5), (x+w+5, y+h+5), (0, 255, 0), 1)
                
        # STEP-7: Saving the boxed image
        cv2.imwrite(os.path.join(os.getcwd(), 'img_boxed', img.split('-')[0] + '-boxed.jpg'), img_obj_boxed)
        
        # STEP-8: Parsing the text 
        parse_txt, word_list, last_word = [], [], ''
        for word in details['text']:
            if word != '':
                word_list.append(word)
                last_word = word
            if (last_word != '' and word == '') or (word == details['text'][-1]):
                parse_txt.append(word_list)
                word_list = []
                
        print('=> done OCR:', count)
        
        parse_txt = list(filter(None, [el for el in parse_txt if len(el) > 1]))
        #=========================================
        # STEP-9 Saving the each file as CSV
        out_file_name = os.path.join(output_dir, img.split('-')[0]+'.csv')
        out_file = open(out_file_name, 'w', newline='')
        csv.writer(out_file, delimiter=" ").writerows(parse_txt)
        out_file.close()
        # print('=> saving csv file to: {}'.format(img.split('-')[0])) TODO
        
    return ''
  

def parse_data(input_dir):
    out_data = {}
    
    files = os.listdir(input_dir)
    files.sort()
    
    for file_name in files:
        file_path = os.path.join(input_dir, file_name)
        file_obj = open(file_path)
        lst = [el.strip() for el in file_obj.readlines()]
        
        lst= clean_lst(lst)
        # Printing the read file name 
        print('File-name:', file_name.split('-')[0])   # TODO
        
        # Printing the extracted raw list 
        print() # TODO 
        pprint(lst)   # TODO
        
        comp_dict = {}
        comp_dict['Company name'] = get_company_name(lst)
        comp_dict['Address'] = get_address(lst)
        comp_dict['Country'] = get_country(comp_dict['Address'])
        comp_dict['Tel'] = get_tel(lst)
        comp_dict['Fax'] = get_fax(lst)
        comp_dict['Website'] = get_website(lst)
        comp_dict['Email'] = get_email(lst, comp_dict['Website'])
        comp_dict['Delegate Name'], comp_dict['Position'] = get_delegate_name(lst)
        
        
        
        # Prints the extracted data 
        print() # TODO
        for keys in comp_dict: # TODO
            print(keys, ':', comp_dict[keys]) # TODO

        print('--- --- --- ---') # TODO 
        
        out_data[file_name] = comp_dict
    return out_data

def get_delegate_name(lst):
    for count, el in enumerate(lst):
        if 'Delegate' in el:
            delegate_index = count
        if ('Te' in el) or ('Ta' in el):
            tel_index = count
    
    delegate, name = [] , []
    
    for el in lst[delegate_index: tel_index]:
        s = el.split('Delegate Name')[-1].strip()
        delegate.append(s.split(', ')[-1].strip())
        name.append(s.split(', ')[0].strip())
    
    # print() # TODO 
    # print(name) # TODO 
    # print(delegate) # TODO 
    # print() # TODO 
    
    return name, delegate


def get_email(lst, websites):
    
    website_index = False
    for count, el in enumerate(lst):
        if 'Email' in el:
            email_index = count
        if 'Website' in el:
            website_index = count
    
    if website_index:
        s = ', '.join(lst[email_index:website_index])
    else:
        s = ', '.join(lst[email_index:])
    
    pattern_1 = re.compile(r'\{d')  # searches for {d to change to @
    pattern_2 = re.compile(r'\(d')  # searches for (d to change to @
    pattern_3 = re.compile(r'fd')   # searches for fd to change to @
    pattern_4 = re.compile(r'fa')   # searches for fa to change to @
    pattern_5 = re.compile(r'id')   # searches for id to change to @
    pattern_6 = re.compile(r'ld')   # searches for id to change to @
    pattern_7 = re.compile(r'td')   # searches for id to change to @
    new_s = pattern_1.sub('@', s)
    new_s = pattern_2.sub('@', new_s)
    new_s = pattern_3.sub('@', new_s)
    new_s = pattern_4.sub('@', new_s)
    new_s = pattern_5.sub('@', new_s)
    new_s = pattern_6.sub('@', new_s)
    new_s = pattern_7.sub('@', new_s)
    
    email_pattern = re.compile(r'([a-zA-Z0-9-\.?]+)@([a-zA-Z0-9-]+\.)([a-zA-Z0-9-]+)(\.[a-zA-Z0-9-]+)?')

    return [match.group() for match in email_pattern.finditer(new_s)]


def get_website(lst):
    website_index = False
    
    for count, el in enumerate(lst):
        if 'Website' in el:
            website_index = count
        
    if website_index:
        s = ', '.join(lst[website_index:])
        
        co_pattern = re.compile(r'www\.\s')
        pattern = re.compile(r'(www\.)([a-zA-Z0-9-]+\.)([a-zA-Z0-9-]+)(\.[a-z]+)?')
        
        s = co_pattern.sub('www.', s)
        return [match.group() for match in pattern.finditer(s)]

    else: 
        s = ''
        return s


def get_fax(lst):
    fax_index = False
    for count, el in enumerate(lst):
        if 'Fax' in el:
            fax_index = count
        if 'Email' in el:
            email_index = count
    
    if fax_index:
        s = ', '.join(lst[fax_index : email_index])
        
        pattern = re.compile(r'(\+\d+\s\d+)(\s\d+)?(\s/\s\d+)?')
        
        return [match.group() for match in pattern.finditer(s)]

    else:
        return ''
    
   
def get_tel(lst): 
    fax_index = False   
    for count, el in enumerate(lst):
        if ('Te' in el) or ('Ta' in el):
            tel_index = count
        else: 
            tel_index: False
        
        if 'Fax' in el:
            fax_index = count
        if 'Email' in el:
            email_index = count
        

    if fax_index:
        s = ', '.join(lst[tel_index : fax_index])
    else:
        s = ', '.join(lst[tel_index : email_index])

    
    pattern = re.compile(r'(\+\d+\]?\s\d+(\s\d+)?)(\sext\.\s\d+)?', re.IGNORECASE)
    sub_pattern = re.compile(r'\]')
    
    s = sub_pattern.sub('1', s)
    return [match.group() for match in pattern.finditer(s)]


def get_country(address):

    pattern = re.compile(r'\d+?')
    lst = pattern.split(address.split(',')[-1])

    return list(filter(None, [el.strip() for el in lst]))[-1]


def get_address(lst):
    for count, el in enumerate(lst):
        if 'Address' in el:
            address_index = count

        if 'Delegate' in el:
            delegate_index = count
    
    return ' '.join(lst[address_index: delegate_index]).split('Address')[-1].strip()

    
def get_company_name(lst):
    
    pattern = re.compile(r'([a-z-A-Z0-9-\s\.\&/])+')
    matches = pattern.finditer(lst[0])
    
    s = ' '.join(''.join([match.group() for match in pattern.finditer(lst[0])]).split(' ')[2:])
    
    return s


def clean_lst(lst):
    if 'CONTACT DETAILS' in lst[0]:
        lst.pop(0)
        
    if not 'Website' in lst[-1]:
        if not 'Email' in lst[-1]:
            lst.pop()
        
    return lst
    
    
# function writes extracted data to dababase excel sheet
def write_to_DB(data_in, out_dir):
    # A1 = No
    # B1 = Company Name 
    # C1 = Country 
    # D1 = Website
    # E1 = Delegate Name 
    # F1 = Position 
    # G1 = Email
    # H1 = Tel 
    # I1 = Fax 
    # J1 = Address
    # K1 = Img-file 
    # L1 = Company introduction / Profile
    
    # TODO read the DB file from the folder 
    file_path = os.path.join(out_dir, 'customer_list-0.xlsx')
    wb_out = openpyxl.load_workbook(file_path)
    
    # TODO get access to the working sheet
    cur_sheet = wb_out['English']
    
    # TODO Check whether rows are empty 
    max_rows = cur_sheet.max_row

    print('The maximum rows: ', max_rows)
    start_row = 2

    # TODO start deliver the data to a certain row from the empty row 
    count = 0
    # print(data_in)
    for name in data_in:
        # for el in data_in[name]:
        # Each data goes to one row in the sheet
        count_abs = count + start_row
        
        cur_sheet.cell(row=count_abs, column=1).value = count_abs - 1
        cur_sheet.cell(row=count_abs, column=2).value = data_in[name]['Company name']
        cur_sheet.cell(row=count_abs, column=3).value = data_in[name]['Country']
        cur_sheet.cell(row=count_abs, column=4).value = ';'.join(data_in[name]['Website'])
        cur_sheet.cell(row=count_abs, column=5).value = '; '.join(data_in[name]['Delegate Name'])
        cur_sheet.cell(row=count_abs, column=6).value = '; '.join(data_in[name]['Position'])
        cur_sheet.cell(row=count_abs, column=7).value = ';'.join(data_in[name]['Email'])
        cur_sheet.cell(row=count_abs, column=8).value = '; '.join(data_in[name]['Tel'])
        cur_sheet.cell(row=count_abs, column=9).value = '; '.join(data_in[name]['Fax'])
        cur_sheet.cell(row=count_abs, column=10).value = data_in[name]['Address']
        cur_sheet.cell(row=count_abs, column=11).value = name
        count += 1

    
    # for count, data in enumerate(data_in):
    #     # Each data goes to one row in the sheet
    #     count_abs = count + start_row
        
        


    
    # TODO after completing save the updated new worksheet 
    wb_out.save(file_path)
    
    
    
    
    print('>Completed writing to DataBase')


# ==========================================================
# ==========================================================
cwd = os.getcwd()

cur_dir = os.path.join(cwd, 'pytesseract')
img_heic_dir = os.path.join(cwd, 'img_heic')
img_jpg_dir = os.path.join(cwd, 'img_jpg')
img_cropp_dir = os.path.join(cwd, 'img_cropped')
out_txt_dir = os.path.join(cwd, 'output-txt')
out_data_dir = os.path.join(cwd, 'output_data')
print(os.path.exists(out_data_dir))

# converting to  jpeg file format from .heic apple image format and cropping Image object
img_cropping(img_heic_dir, img_cropp_dir)             # TODO

# reading the text from cropped files
print('ocr processing...')
img_to_text(img_cropp_dir, out_txt_dir)               # TODO
print('--- --- --- OCR-done --- --- ---')


print('--- --- --- Parse data... -- --- ')  
DATA = parse_data(out_txt_dir)                          # TODO 
# print(DATA) # Printing the clean data                 # TODO
print('--- --- --- Parse done! --- --- --- ') 
print() 

print('writing to Excel DB...')
write_to_DB(DATA, out_data_dir)                         # TODO
print('--- --- --- Writing done! --- --- --- ')         #TODO
print()                                                 #TODO

print('--- --- --- ALL TASKS COMPLETED --- --- --- ')